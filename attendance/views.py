import base64
import io
import json
from datetime import date, datetime, timedelta
from django.shortcuts import render, redirect
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse, HttpResponse
from django.core.files.base import ContentFile
from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference, BarChart
from openpyxl.styles import Font
from .forms import EmployeeSignUpForm, LoginForm
from .models import EmployeeProfile, Attendance

# try/import face_recognition but fail gracefully if not installed
try:
    import face_recognition
    import numpy as np
except Exception as exc:
    face_recognition = None

def home_view(request):
    return redirect('login')

from django.contrib.auth.decorators import login_required
from django.db import IntegrityError

def register_view(request):
    # CASE 1: Logged-in user → complete profile only
    if request.user.is_authenticated:
        user = request.user

        if request.method == 'POST':
            employee_id = request.POST.get('employee_id')
            department = request.POST.get('department')
            face_image_data = request.POST.get('face_image_data')

            profile, created = EmployeeProfile.objects.get_or_create(
                user=user,
                defaults={
                    'employee_id': employee_id,
                    'department': department
                }
            )

            if not created:
                profile.employee_id = employee_id
                profile.department = department

            if face_image_data:
                header, encoded = face_image_data.split(',', 1)
                image_data = base64.b64decode(encoded)
                file_name = f'face_{employee_id}.png'
                profile.face_image.save(file_name, ContentFile(image_data), save=True)

                if face_recognition:
                    image = face_recognition.load_image_file(profile.face_image.path)
                    encodings = face_recognition.face_encodings(image)
                    if encodings:
                        profile.face_encoding = json.dumps(encodings[0].tolist())

            profile.save()
            return redirect('dashboard')

        return render(request, 'register.html', {'profile_only': True})

    # CASE 2: New user signup
    if request.method == 'POST':
        form = EmployeeSignUpForm(request.POST)
        if form.is_valid():
            try:
                user = User.objects.create_user(
                    username=form.cleaned_data['username'],
                    password=form.cleaned_data['password'],
                    first_name=form.cleaned_data['first_name'],
                    last_name=form.cleaned_data['last_name'],
                    email=form.cleaned_data.get('email', '')
                )
            except IntegrityError:
                return render(request, 'register.html', {
                    'form': form,
                    'error': 'Username already exists. Please login.'
                })

            EmployeeProfile.objects.create(
                user=user,
                employee_id=form.cleaned_data['employee_id'],
                department=form.cleaned_data.get('department', '')
            )

            login(request, user)
            return redirect('dashboard')

    else:
        form = EmployeeSignUpForm()

    return render(request, 'register.html', {'form': form})

def register_face_view(request):
    # simplified separate endpoint to register only face for existing user
    if request.method == 'POST':
        username = request.POST.get('username')
        face_data = request.POST.get('face_data')
        try:
            user = User.objects.get(username=username)
            profile = EmployeeProfile.objects.get(user=user)
        except Exception:
            return JsonResponse({'status': 'error', 'message': 'User not found'}, status=404)
        header, encoded = face_data.split(',', 1)
        image_data = base64.b64decode(encoded)
        file_name = f'face_{profile.employee_id}.png'
        profile.face_image.save(file_name, ContentFile(image_data), save=True)

        if face_recognition:
            image = face_recognition.load_image_file(profile.face_image.path)
            encodings = face_recognition.face_encodings(image)
            if encodings:
                profile.face_encoding = json.dumps(encodings[0].tolist())
                profile.save()
        return JsonResponse({'status': 'ok'})
    return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=400)

def login_view(request):
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            user = authenticate(request, username=username, password=password)
            if user:
                login(request, user)
                return redirect('dashboard')
            else:
                return render(request, 'login.html', {'form': form, 'error': 'Invalid credentials'})
    else:
        form = LoginForm()
    return render(request, 'login.html', {'form': form})

def logout_view(request):
    logout(request)
    return redirect('login')

def dashboard_view(request):
    if not request.user.is_authenticated:
        return redirect('login')

    try:
        profile = EmployeeProfile.objects.get(user=request.user)
    except EmployeeProfile.DoesNotExist:
        return redirect('register')

    attendance_entries = []

    if profile:
        today = date.today()
        number_of_days = 7

        # Generate last N dates (including today)
        recent_dates = [today - timedelta(days=day) for day in range(number_of_days)]

        # Fetch existing attendance records for the employee
        existing_records = Attendance.objects.filter(
            employee=profile,
            date__in=recent_dates
        )

        # Map records by date for fast lookup
        attendance_by_date = {
            record.date: record for record in existing_records
        }

        # Build final attendance list (Present + Absent)
        for current_date in recent_dates:
            if current_date in attendance_by_date:
                record = attendance_by_date[current_date]
                attendance_entries.append({
                    'date': current_date,
                    'check_in_time': record.check_in_time,
                    'status': 'Present'
                })
            else:
                attendance_entries.append({
                    'date': current_date,
                    'check_in_time': None,
                    'status': 'Absent'
                })

    return render(request, 'dashboard.html', {
        'profile': profile,
        'attendance_entries': attendance_entries
    })


@csrf_exempt
def recognize_view(request):
    'API endpoint to accept webcam capture and mark attendance if face matches'
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'POST required'}, status=400)
    if not face_recognition:
        return JsonResponse({'status': 'error', 'message': 'face_recognition library not installed'}, status=500)

    data = json.loads(request.body.decode('utf-8'))
    image_data = data.get('image')
    if not image_data:
        return JsonResponse({'status': 'error', 'message': 'No image provided'}, status=400)
    header, encoded = image_data.split(',', 1)
    image_bytes = base64.b64decode(encoded)
    image_stream = io.BytesIO(image_bytes)
    image = face_recognition.load_image_file(image_stream)
    faces_encodings = face_recognition.face_encodings(image)
    if not faces_encodings:
        return JsonResponse({'status': 'error', 'message': 'No face detected'}, status=400)

    found_encoding = faces_encodings[0]
    # gather all stored employee encodings
    all_profiles = EmployeeProfile.objects.exclude(face_encoding__isnull=True).exclude(face_encoding__exact='')
    known_encodings = []
    known_employee_ids = []
    for profile in all_profiles:
        try:
            encoding_list = json.loads(profile.face_encoding)
            known_encodings.append(encoding_list)
            known_employee_ids.append(profile.employee_id)
        except Exception:
            continue

    import numpy as np
    matches = face_recognition.compare_faces([np.array(e) for e in known_encodings], np.array(found_encoding), tolerance=0.5)
    if True in matches:
        matched_index = matches.index(True)
        matched_employee_id = known_employee_ids[matched_index]
        try:
            matched_profile = EmployeeProfile.objects.get(employee_id=matched_employee_id)
        except EmployeeProfile.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Matched profile not found'}, status=500)

        today = date.today()
        now_time = datetime.now().time()
        attendance_record, created = Attendance.objects.get_or_create(employee=matched_profile, date=today,
                                                                      defaults={'check_in_time': now_time, 'status': 'Present'})
        if not created:
            # already exists, update check_in_time only if empty
            if not attendance_record.check_in_time:
                attendance_record.check_in_time = now_time
                attendance_record.save()
        return JsonResponse({'status': 'ok', 'employee': matched_profile.employee_id, 'first_name': matched_profile.user.first_name})
    else:
        return JsonResponse({'status': 'error', 'message': 'No match found'}, status=404)

def attendance_list_view(request):
    if not request.user.is_authenticated:
        return redirect('login')

    today = date.today()
    number_of_days = 30
    recent_dates = [today - timedelta(days=i) for i in range(number_of_days)]

    attendance_entries = []

    if request.user.is_staff:
        # Admin can see all employees
        employees = EmployeeProfile.objects.select_related('user').all()
    else:
        # Normal employee sees only own data
        try:
            employees = [EmployeeProfile.objects.select_related('user').get(user=request.user)]
        except EmployeeProfile.DoesNotExist:
            employees = []

    for employee in employees:
        records = Attendance.objects.filter(
            employee=employee,
            date__in=recent_dates
        )

        record_map = {record.date: record for record in records}

        for current_date in recent_dates:
            if current_date in record_map:
                record = record_map[current_date]
                attendance_entries.append({
                    'employee': employee,
                    'date': current_date,
                    'check_in_time': record.check_in_time,
                    'status': 'Present'
                })
            else:
                attendance_entries.append({
                    'employee': employee,
                    'date': current_date,
                    'check_in_time': None,
                    'status': 'Absent'
                })

    return render(request, 'attendance_list.html', {
        'entries': attendance_entries,
        'employees': employees if request.user.is_staff else []  # for admin dropdown
    })

@staff_member_required
def download_attendance_excel(request):
    employee_id = request.GET.get('employee_id')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if not employee_id or not start_date or not end_date:
        return HttpResponse('Employee and date range are required')

    employee = EmployeeProfile.objects.select_related('user').get(id=employee_id)

    start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Attendance Report'

    headers = [
        'Employee Name',
        'Employee ID',
        'Department',
        'Date',
        'Check-in Time',
        'Status',
        'Status Value'
    ]
    sheet.append(headers)
    for col in range(1, len(headers) + 1):
        sheet.cell(row=1, column=col).font = Font(bold=True)

    present_count = 0
    absent_count = 0

    total_days = (end_date - start_date).days + 1
    all_dates = [start_date + timedelta(days=i) for i in range(total_days)]

    records = Attendance.objects.filter(
        employee=employee,
        date__range=[start_date, end_date]
    )
    record_map = {record.date: record for record in records}

    for current_date in all_dates:
        if current_date in record_map:
            record = record_map[current_date]
            status = 'Present'
            status_value = 1
            present_count += 1
            check_in = record.check_in_time.strftime('%H:%M:%S') if record.check_in_time else ''
        else:
            status = 'Absent'
            status_value = 0
            absent_count += 1
            check_in = ''

        sheet.append([
            employee.user.get_full_name(),
            employee.employee_id,
            employee.department,
            current_date.strftime('%Y-%m-%d'),
            check_in,
            status,
            status_value
        ])

    last_row = sheet.max_row

    summary_row = last_row + 3
    sheet[f'A{summary_row}'] = 'Summary'
    sheet[f'A{summary_row}'].font = Font(bold=True)

    sheet.append(['Present', present_count])
    sheet.append(['Absent', absent_count])

    summary_start = summary_row + 1

    pie = PieChart()
    labels = Reference(sheet, min_col=1, min_row=summary_start, max_row=summary_start + 1)
    data = Reference(sheet, min_col=2, min_row=summary_start, max_row=summary_start + 1)

    pie.add_data(data, titles_from_data=False)
    pie.set_categories(labels)
    pie.title = 'Attendance Distribution'

    sheet.add_chart(pie, 'I2')

    bar = BarChart()
    bar.title = 'Date-wise Attendance'
    bar.y_axis.title = 'Present (1) / Absent (0)'
    bar.x_axis.title = 'Date'

    values = Reference(sheet, min_col=7, min_row=2, max_row=last_row)
    categories = Reference(sheet, min_col=4, min_row=2, max_row=last_row)

    bar.add_data(values, titles_from_data=False)
    bar.set_categories(categories)

    sheet.add_chart(bar, 'I20')

    filename = f'{employee.employee_id}_{start_date}_to_{end_date}.xlsx'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'

    workbook.save(response)
    return response
